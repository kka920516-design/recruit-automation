import streamlit as st
import pandas as pd
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="招募日報自動化", page_icon="📋", layout="centered")

# ============================================================
# 密碼保護
# ============================================================
def check_password():
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if not st.session_state.authenticated:
        st.title("🔒 招募日報自動化")
        st.caption("請輸入密碼以繼續")
        pwd = st.text_input("密碼", type="password")
        if st.button("登入", use_container_width=True):
            if pwd == st.secrets["APP_PASSWORD"]:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("密碼錯誤，請再試一次")
        return False
    return True

if not check_password():
    st.stop()
# ============================================================

st.title("📋 招募日報自動化")
st.caption("上傳招募報表，自動合併並產生 LINE 訊息")

COLUMN_MAPPING = {
    "更新時間": "更新日期",
    "Recruiter": "招募人員",
    "姓名": "人選姓名",
    "職缺": "職務名稱",
    "聯繫說明": "聯繫註記",
    "備註": "聯繫註記",
    "錄取\n與否": "錄取狀態",
    "錄取#": "錄取狀態備註",
    "到職W#": "到職W#",
    "Final \nStatus": "Final Status",
    "客戶\n轉掛": "客戶轉掛",
    "最後修\n改時間": "最後修改時間",
}

OUTPUT_COLUMNS = [
    "更新日期", "招募人員", "人選姓名", "客戶", "職務名稱",
    "聯繫註記", "聯絡日期", "一面日期", "履歷日期\n(客戶)",
    "履歷狀態", "二面日期", "面試狀態", "錄取時間", "錄取狀態",
    "聘書日期", "期望薪資", "在職情況", "到職日", "最後修改時間", "來源檔案"
]

DATE_COLS = [
    "更新日期", "聯絡日期", "一面日期", "履歷日期\n(客戶)",
    "二面日期", "錄取時間", "聘書日期", "到職日", "最後修改時間"
]

COL_WIDTHS = {
    "更新日期": 12, "招募人員": 10, "人選姓名": 10, "客戶": 16,
    "職務名稱": 22, "聯繫註記": 50, "聯絡日期": 12, "一面日期": 12,
    "履歷日期\n(客戶)": 12, "履歷狀態": 10, "二面日期": 12,
    "面試狀態": 10, "錄取時間": 12, "錄取狀態": 10, "聘書日期": 12,
    "期望薪資": 12, "在職情況": 10, "到職日": 12,
    "最後修改時間": 16, "來源檔案": 24,
}

BOLD_COLS = {"職務名稱", "人選姓名"}

def parse_date(val):
    if pd.isna(val) or str(val).strip() == "":
        return ""
    s = str(val).strip()
    if "月" in s and "日" in s and len(s) <= 6:
        try:
            year = datetime.today().year
            dt = pd.to_datetime(f"{year}年{s}", format="%Y年%m月%d日", errors="coerce")
            if pd.notna(dt):
                return dt.strftime("%Y/%m/%d")
        except:
            pass
    try:
        dt = pd.to_datetime(s, errors="coerce")
        if pd.notna(dt):
            return dt.strftime("%Y/%m/%d")
    except:
        pass
    return s

def process_file(file_obj, filename):
    xl = pd.ExcelFile(file_obj)
    target_sheets = [s for s in xl.sheet_names if "招募狀態" in s]
    if not target_sheets:
        return None
    dfs = []
    for sname in target_sheets:
        df = pd.read_excel(file_obj, sheet_name=sname, dtype=str)
        df = df.dropna(how="all")
        df = df[[c for c in df.columns if pd.notna(c) and str(c).strip() != ""]]
        df = df.rename(columns=COLUMN_MAPPING)
        df["來源檔案"] = filename
        dfs.append(df)
    return pd.concat(dfs, ignore_index=True, sort=False) if len(dfs) > 1 else dfs[0]

def style_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    header_fill  = PatternFill("solid", start_color="1F4E79")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    no_fill = PatternFill(fill_type=None)
    thin   = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for col in range(1, ws.max_column + 1):
        col_name = headers[col - 1]
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS.get(str(col_name), 14)
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 45
        for col in range(1, ws.max_column + 1):
            col_name = headers[col - 1]
            cell = ws.cell(row=row, column=col)
            cell.fill = no_fill
            cell.border = border
            if str(col_name) == "聯繫註記":
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.font = Font(name="Arial", size=10, bold=str(col_name) in BOLD_COLS)
                cell.alignment = Alignment(vertical="center", wrap_text=False)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 32

# UI
target_date = st.date_input("篩選日期", value=datetime.today())
uploaded_files = st.file_uploader("上傳招募報表（可多選）", type=["xlsx", "xls"], accept_multiple_files=True)

if st.button("🚀 開始執行", use_container_width=True, type="primary"):
    if not uploaded_files:
        st.error("請先上傳報表！")
    else:
        target_date_str = target_date.strftime("%Y/%m/%d")
        with st.spinner("處理中..."):
            try:
                all_dfs = []
                for f in uploaded_files:
                    df = process_file(f, f.name)
                    if df is None:
                        st.warning(f"⚠️ {f.name} 找不到招募狀態 sheet，已跳過")
                        continue
                    for col in DATE_COLS:
                        if col in df.columns:
                            df[col] = df[col].apply(parse_date)
                    if "更新日期" in df.columns:
                        df = df[df["更新日期"] == target_date_str]
                    all_dfs.append(df)

                if not all_dfs:
                    st.error("沒有找到符合日期的資料，請確認報表內容")
                else:
                    merged = pd.concat(all_dfs, ignore_index=True, sort=False)
                    ordered = [c for c in OUTPUT_COLUMNS if c in merged.columns]
                    extras  = [c for c in merged.columns if c not in ordered]
                    merged  = merged[ordered + extras]

                    contact_col = "聯絡日期"
                    if contact_col in merged.columns:
                        sheet1 = merged[merged[contact_col] == target_date_str].copy()
                        sheet2 = merged[merged[contact_col] != target_date_str].copy()
                    else:
                        sheet1, sheet2 = merged.copy(), pd.DataFrame()

                    resume_col = "履歷日期\n(客戶)"
                    delivery = int((merged[resume_col] == target_date_str).sum()) if resume_col in merged.columns else len(merged)
                    line_msg = f"Hi all, 這裡跟招募交付們確認今日收到履歷共計 {delivery} 份"

                    # 輸出 Excel
                    buf = io.BytesIO()
                    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                        sheet1.to_excel(writer, sheet_name="今日新增", index=False)
                        if not sheet2.empty:
                            sheet2.to_excel(writer, sheet_name="歷史更新", index=False)
                    buf.seek(0)
                    wb = load_workbook(buf)
                    style_sheet(wb, "今日新增")
                    if "歷史更新" in wb.sheetnames:
                        style_sheet(wb, "歷史更新")
                    out = io.BytesIO()
                    wb.save(out)
                    out.seek(0)

                    # 顯示結果
                    st.success("✅ 執行完成！")
                    col1, col2, col3 = st.columns(3)
                    col1.metric("今日新增", f"{len(sheet1)} 筆")
                    col2.metric("歷史更新", f"{len(sheet2)} 筆")
                    col3.metric("📬 履歷交付量", f"{delivery} 份")

                    st.info(f"💬 LINE 訊息：\n\n**{line_msg}**")
                    st.code(line_msg, language=None)

                    fname = f"招募狀態合併_{target_date.strftime('%Y%m%d')}.xlsx"
                    st.download_button("⬇️ 下載 Excel", data=out, file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)

            except Exception as e:
                st.error(f"執行失敗：{e}")
